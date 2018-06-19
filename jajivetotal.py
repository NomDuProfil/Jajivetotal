import requests
import json
import sys, os
from openpyxl import Workbook
import dns.resolver

#https://api.passivetotal.org/api/docs/#api-Enrichment-GetV2EnrichmentSubdomains

credentials = {
  "USERNAME": ['email1', 
  		'email2', 
  		'email3'],

  "API_KEY": ['email1_key',
              'email2_key', 
              'email3_key'],
}

class domainsClass(object):
	def __init__(self):
		self._name = ""
		self._subdomains = []

	@property
	def name(self):
		return self._name

	@name.setter
	def name(self, n):
		self._name = n
		return self._name

	@property
	def subdomains(self):
		return self._subdomains

	@subdomains.setter
	def subdomains(self, s):
		self._subdomains = self._subdomains+s
		return self._subdomains

class ipClass(object):
	def __init__(self):
		self._ip = ""
		self._domains = []

	@property
	def ip(self):
		return self._ip

	@ip.setter
	def ip(self, i):
		self._ip = i
		return self._ip

	@property
	def domains(self):
		return self._domains

	@domains.setter
	def domains(self, d):
		self._domains = self._domains+[d]
		return self._domains

activeaccount = 0
totaldomains = []

def generatefiles():

	print "Generate files..."

	global totaldomains

	totalipdomains = []

	for current in totaldomains:
		iptmp = getip(current.name)
		checkip = checkippresent(iptmp, totalipdomains)
		if checkip == -1:
			newip = ipClass()
			newip.ip = iptmp
			newip.domains = current.name
			totalipdomains.append(newip)
		else:
			totalipdomains[checkip].domains = current.name
		for currentsub in current.subdomains:
			iptmp = getip(currentsub+'.'+current.name)
			checkip = checkippresent(iptmp, totalipdomains)
			if checkip == -1:
				newip = ipClass()
				newip.ip = iptmp
				newip.domains = currentsub+'.'+current.name
				totalipdomains.append(newip)
			else:
				totalipdomains[checkip].domains = currentsub+'.'+current.name

	#### EXCEL FILE ####

	wb = Workbook()
	ws = wb.active
	ws['A1'] = "Domains"
	ws['B1'] = "IP"
	countline = 2
	for current in totalipdomains:
		ws["B"+str(countline)] = current.ip
		domainstring = ""
		for currentdomain in current.domains:
			if domainstring == "":
				domainstring = currentdomain
			else:
				domainstring = domainstring+"~"+currentdomain
		ws["A"+str(countline)] = domainstring
		countline += 1

	wb.save(domainsearch.replace(".", "_")+".xlsx")

	print "XLSX done..."

	#### CSV FILE ####

	csv = open(domainsearch.replace(".", "_"), "w") 

	columnName = "Domains, IP\n"
	csv.write(columnName)
	for current in totalipdomains:
		totalstring = ""
		totalstring = current.ip+", "
		for currentdomain in current.domains:
			if domainstring == "":
				domainstring = currentdomain
			else:
				domainstring = domainstring+"~"+currentdomain
		totalstring = totalstring+domainstring+"\n"
		csv.write(totalstring)
	csv.close()
	print "CSV done..."

def getip(domain):
	ip = ""
	try:
		myResolver = dns.resolver.Resolver()
		myAnswers = myResolver.query(domain, "A")
		for rdata in myAnswers:
			if ip != "":
				ip = ip+"~"+str(rdata)
			else:
				ip = str(rdata)
	except dns.resolver.NXDOMAIN:
		ip="NO_RESOLUTION"
		country="UNKNOWN"
	except dns.resolver.NoAnswer:
		ip="NO_RESOLUTION"
		country="UNKNOWN"
	except dns.exception.Timeout:
		ip="TIMEOUT"
		country="UNKNOWN"
	except dns.resolver.NoNameservers:
		ip="NO_NAME_SERVER"
		country="UNKNOWN"
	return ip

def checkippresent(ip, struct):
	compt = 0
	for current in struct:
		if current.ip == ip:
			return compt
	return -1

def checkerror(result):
	global activeaccount
	if "message" in result:
		if "quota_exceeded" in result["message"]:
		    	if activeaccount == len(credentials["USERNAME"])-1:
		    		print "Quota exceeded for "+credentials["USERNAME"][activeaccount]
		    		print "No account available and quota exceeded"
		    		generatefiles()
		    		sys.exit(0)
		    	print "Quota exceeded, change account "+credentials["USERNAME"][activeaccount]+" for "+credentials["USERNAME"][activeaccount+1]
		    	activeaccount += +1
		    	return 0
		elif "invalid credentials" in result["message"]:
			print "Invalid credentials for "+credentials["USERNAME"][activeaccount]
			generatefiles()
			sys.exit(0)
	else:
		return result

def passivetotal_get(path, query, field=None):
    global activeaccount
    auth = (credentials["USERNAME"][activeaccount], credentials["API_KEY"][activeaccount])
    url = "https://api.passivetotal.org" + path
    if field != None:
    	data = {'query': query, 'field': field}
    else:
    	data = {'query': query}
    response = requests.get(url, auth=auth, json=data)
    while True:
    	result = checkerror(response.json())
    	if result == 0:
    		auth = (credentials["USERNAME"][activeaccount], credentials["API_KEY"][activeaccount])
    		response = requests.get(url, auth=auth, json=data)
    	else:
    		break
    return result

def getmailfordomain(domain):
	result = passivetotal_get("/v2/whois/search", domain, field="domain")
	if result["results"][0]["admin"]["email"]:
		return result["results"][0]["admin"]["email"]
	return ""

def getdomainsforemail(email):
	domains = []
	result = passivetotal_get("/v2/whois/search", mail, field="email")
	for current in result["results"]:
		domains.append(current["domain"])
	return domains

def getsubdomainsfordomain(domain):
	subdomains = []
	result = passivetotal_get("/v2/enrichment/subdomains", domain)
	for current in result["subdomains"]:
		subdomains.append(current)
	return subdomains

if len(sys.argv) < 2:
    print "usage: python jajivetotal.py domainsearch"
    sys.exit(0)

print "Active account :"+credentials["USERNAME"][activeaccount]

domainsearch = sys.argv[1]


mail = getmailfordomain(domainsearch)

if mail != "":
	domains = getdomainsforemail(mail)
else:
	domains = [domainsearch]

for current in domains:
	newdomain = domainsClass()
	newdomain.name = current
	newdomain.subdomains = getsubdomainsfordomain(current)
	totaldomains.append(newdomain)

generatefiles()
